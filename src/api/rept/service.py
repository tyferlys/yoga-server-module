from io import BytesIO

import pandas as pd
from datetime import datetime

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from src.database.config import sync_engine
from openpyxl import Workbook

class ReptService:
    def get_rept_general(self, begin_date: datetime, end_date: datetime):
        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', None)
        pd.set_option('display.max_colwidth', None)

        df = pd.read_sql_query(f"""
            select	rp.answer as answer, 
                    yp_de.title_russian as default_answer,
                    yp1.title_russian as is_right_top1,
                    yp5.title_russian as is_right_top5,
                    yp_ra.title_russian as right_answer_system, 
                    rp.right_answer_sanskrit, 
                    rp.right_transliteration,
                    rp.right_answer_russian,
                    rp.right_answer_russian_interpretation,
                    rp.created_at
            from t_result_predictions rp
            left join t_yoga_poses yp1 on is_right_top1 = yp1.id
            left join t_yoga_poses yp5 on is_right_top5 = yp5.id
            left join t_yoga_poses yp_ra on right_answer_system = yp_ra.id
            left join t_yoga_poses yp_de on (string_to_array(trim(both '[]' from answer), ', '))[1]::int = yp_de.id
            where created_at between %(begin_date)s and %(end_date)s
        """, con=sync_engine, params={'begin_date': begin_date, 'end_date': end_date})

        df_is_right_top1 = df[df["is_right_top1"].notna()]
        df_is_right_top5 = df[df["is_right_top5"].notna()]
        df_is_wrong = df[df["right_answer_system"].notna()]
        df_is_not_exists= df[
            (df["right_answer_sanskrit"].notna()) |
            (df["right_transliteration"].notna()) |
            (df["right_answer_russian"].notna()) |
            (df["right_answer_russian_interpretation"].notna())
        ]
        df_empty = df[
            (df["right_answer_sanskrit"].isna()) &
            (df["right_transliteration"].isna()) &
            (df["right_answer_russian"].isna()) &
            (df["right_answer_russian_interpretation"].isna()) &
            (df["is_right_top1"].isna()) &
            (df["is_right_top5"].isna()) &
            (df["right_answer_system"].isna())
        ]

        # количество предсказаний
        df_count = len(df)
        df_is_right_top1_count = len(df_is_right_top1)
        df_is_right_top5_count = len(df_is_right_top5)
        df_is_wrong_count = len(df_is_wrong)
        df_is_not_exists_count = len(df_is_not_exists)
        df_empty_count = len(df_empty)

        # список асан которые определяются правильно
        df_is_right_top1_general = df_is_right_top1.groupby(["is_right_top1"]).size().reset_index(name='count').sort_values('count', ascending=False)

        # список асан которые определяются не очень правильно
        df_is_right_top5_general = df_is_right_top5.groupby(["is_right_top5"]).size().reset_index(name='count').sort_values('count', ascending=False)

        # список асан которые определяются неправильно
        df_is_wrong_count_general = df_is_wrong.groupby(["right_answer_system"]).size().reset_index(name='count').sort_values('count', ascending=False)

        # список асан которые пустые
        df_empty_count_general = df_empty.groupby(["default_answer"]).size().reset_index(
            name='count').sort_values('count', ascending=False)

        df_asana_general = df_is_right_top1_general \
            .merge(df_is_right_top5_general, left_on="is_right_top1", right_on="is_right_top5", how="outer", suffixes=('', '_top5')) \
            .merge(df_is_wrong_count_general, left_on="is_right_top1", right_on="right_answer_system", how="outer", suffixes=('', '_wrong')) \
            .merge(df_empty_count_general, left_on="is_right_top1", right_on="default_answer", how="outer", suffixes=('', '_empty')) \
            .fillna(0)
        df_asana_general["all_count"] = df_asana_general["count"] + df_asana_general["count_top5"] + df_asana_general["count_wrong"] + df_asana_general["count_empty"]
        df_asana_general = df_asana_general.sort_values("all_count", ascending=False)

        # отсутствующие асаны
        df_asans_not_exists = df_is_not_exists[["right_answer_sanskrit", "right_transliteration", "right_answer_russian", "right_answer_russian_interpretation"]] \
            .fillna("Отсутствует") \
            .groupby([
                "right_answer_sanskrit",
                "right_transliteration",
                "right_answer_russian",
                "right_answer_russian_interpretation"
            ]).size().reset_index(name='count').sort_values('count', ascending=False)


        wb = Workbook()
        ws = wb.active
        ws.title = "Общий отчет"
        bold_font = Font(bold=True)

        ws.cell(row=1, column=1, value="Общее количество предсказаний")
        ws.cell(row=1, column=2, value="Количество верных top1 предсказаний")
        ws.cell(row=1, column=3, value="Количество верных top5 предсказаний")
        ws.cell(row=1, column=4, value="Количество неверных предсказаний")
        ws.cell(row=1, column=5, value="Количество предсказаний в которых верной асаны нет в системе")
        ws.cell(row=1, column=6, value="Количество не заполненных предсказаний")
        for cell in ws[1]:
            cell.font = bold_font

        ws.cell(row=2, column=1, value=df_count)
        ws.cell(row=2, column=2, value=df_is_right_top1_count)
        ws.cell(row=2, column=3, value=df_is_right_top5_count)
        ws.cell(row=2, column=4, value=df_is_wrong_count)
        ws.cell(row=2, column=5, value=df_is_not_exists_count)
        ws.cell(row=2, column=6, value=df_empty_count)

        ws.cell(row=4, column=1, value="Список асан, которые были определены в предсказаниях")
        for cell in ws[4]:
            cell.font = bold_font
        ws.cell(row=5, column=1, value="Название асаны")
        ws.cell(row=5, column=2, value="Общее количество предсказаний")
        ws.cell(row=5, column=3, value="Количество верных top1 предсказаний")
        ws.cell(row=5, column=4, value="Количество верных top5 предсказаний")
        ws.cell(row=5, column=5, value="Количество неверных предсказаний")
        ws.cell(row=5, column=6, value="Количество не заполненных предсказаний")
        for cell in ws[5]:
            cell.font = bold_font

        for i in range(0, len(df_asana_general)):
            not_none_title = list(
                filter(lambda x: x is not None and x != 0, [
                    df_asana_general["is_right_top1"].iloc[i],
                    df_asana_general["is_right_top5"].iloc[i],
                    df_asana_general["right_answer_system"].iloc[i],
                    df_asana_general["default_answer"].iloc[i]
                ])
            )[0]

            ws.cell(row=i + 6, column=1, value=not_none_title)
            ws.cell(row=i + 6, column=2, value=df_asana_general["all_count"].iloc[i])
            ws.cell(row=i + 6, column=3, value=df_asana_general["count"].iloc[i])
            ws.cell(row=i + 6, column=4, value=df_asana_general["count_top5"].iloc[i])
            ws.cell(row=i + 6, column=5, value=df_asana_general["count_wrong"].iloc[i])
            ws.cell(row=i + 6, column=6, value=df_asana_general["count_empty"].iloc[i])

        next_row = len(df_asana_general) + 6
        ws.cell(row=next_row + 1, column=1, value="Информация об асанах, которых нет в системе")
        for cell in ws[next_row + 1]:
            cell.font = bold_font

        ws.cell(row=next_row + 2, column=1, value="Название асаны на Санскрите")
        ws.cell(row=next_row + 2, column=2, value="Транслитерация")
        ws.cell(row=next_row + 2, column=3, value="Название асаны на Русском")
        ws.cell(row=next_row + 2, column=4, value="Перевод названия асаны")
        ws.cell(row=next_row + 2, column=5, value="Количество предсказаний с этой асанами")
        for cell in ws[next_row + 2]:
            cell.font = bold_font

        for i in range(0, len(df_asans_not_exists)):
            ws.cell(row=i + next_row + 3, column=1, value=df_asans_not_exists["right_answer_sanskrit"].iloc[i])
            ws.cell(row=i + next_row + 3, column=2, value=df_asans_not_exists["right_transliteration"].iloc[i])
            ws.cell(row=i + next_row + 3, column=3, value=df_asans_not_exists["right_answer_russian"].iloc[i])
            ws.cell(row=i + next_row + 3, column=4, value=df_asans_not_exists["right_answer_russian_interpretation"].iloc[i])
            ws.cell(row=i + next_row + 3, column=5, value=df_asans_not_exists["count"].iloc[i])

        for col_cells in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max_len + 2

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        return stream


if __name__ == '__main__':
    test = ReptService()
    test.get_rept_general(datetime(2025, 1, 1), datetime(2025, 4, 29))